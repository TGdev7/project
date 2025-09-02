from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from admin_panel.permissions import IsAdminUser

from users.models import *




#############################report generation code###################################
import csv
import io
from datetime import datetime
from django.http import HttpResponse
from django.db.models import Q
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import json
from rest_framework.permissions import AllowAny     

class GenericReportDownloadView(APIView):
    """
    Generic API view for downloading reports from Django model data.
    
    Supports CSV, Excel, and JSON formats with filtering capabilities.
    """
    permission_classes = [AllowAny]
    
    # Override these in your subclass
    model = None
    fields = []  # List of field names to include in report
    field_labels = {}  # Dict mapping field names to display labels
    queryset = None
    
    def get_queryset(self):
        """Get the base queryset for the report."""
        if self.queryset is not None:
            return self.queryset
        if self.model is not None:
            return self.model.objects.all()
        raise NotImplementedError("Must define 'model' or 'queryset' attribute")
    
    def get_fields(self):
        """Get the fields to include in the report."""
        if not self.fields:
            raise NotImplementedError("Must define 'fields' attribute")
        return self.fields
    
    def get_field_labels(self):
        """Get display labels for fields."""
        labels = {}
        for field in self.get_fields():
            labels[field] = self.field_labels.get(field, field.replace('_', ' ').title())
        return labels
    
    def apply_filters(self, queryset, filters):
        """Apply filters to the queryset."""
        if not filters:
            return queryset
        
        q_objects = Q()
        
        for field, value in filters.items():
            if field.endswith('__gte'):
                q_objects &= Q(**{field: value})
            elif field.endswith('__lte'):
                q_objects &= Q(**{field: value})
            elif field.endswith('__contains'):
                q_objects &= Q(**{field: value})
            elif field.endswith('__icontains'):
                q_objects &= Q(**{field: value})
            elif field.endswith('__in'):
                if isinstance(value, str):
                    value = value.split(',')
                q_objects &= Q(**{field: value})
            else:
                q_objects &= Q(**{field: value})
        
        return queryset.filter(q_objects)
    
    def get_field_value(self, obj, field):
        """Get the value of a field from an object, handling nested fields."""
        if '__' in field:
            # Handle foreign key relationships
            parts = field.split('__')
            value = obj
            for part in parts:
                if hasattr(value, part):
                    value = getattr(value, part)
                else:
                    return None
            return value
        else:
            return getattr(obj, field, None)
    
    def format_value(self, value):
        """Format a value for display in the report."""
        if value is None:
            return ''
        elif isinstance(value, bool):
            return 'Yes' if value else 'No'
        elif hasattr(value, 'strftime'):
            return value.strftime('%Y-%m-%d %H:%M:%S')
        else:
            return str(value)
    
    def generate_csv(self, data, field_labels):
        """Generate CSV report."""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        headers = [field_labels[field] for field in self.get_fields()]
        writer.writerow(headers)
        
        # Write data rows
        for obj in data:
            row = []
            for field in self.get_fields():
                value = self.get_field_value(obj, field)
                row.append(self.format_value(value))
            writer.writerow(row)
        
        output.seek(0)
        return output.getvalue()
    
    def generate_excel(self, data, field_labels):
        """Generate Excel report."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"
        
        # Style for headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Write headers
        headers = [field_labels[field] for field in self.get_fields()]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Write data
        for row_idx, obj in enumerate(data, 2):
            for col_idx, field in enumerate(self.get_fields(), 1):
                value = self.get_field_value(obj, field)
                formatted_value = self.format_value(value)
                ws.cell(row=row_idx, column=col_idx, value=formatted_value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
   
    def get(self, request):
        """Handle GET request for downloading reports."""
        try:
            # Get query parameters
            format_type = request.query_params.get('format', 'csv').lower()
            filename = request.query_params.get('filename', 'report')
            
            # Get filters from query parameters
            filters = {}
            for key, value in request.query_params.items():
                if key not in ['format', 'filename']:
                    filters[key] = value
            
            # Get and filter data
            queryset = self.get_queryset()
            queryset = self.apply_filters(queryset, filters)
            
            # Select related fields to optimize queries
            related_fields = [field.split('__')[0] for field in self.get_fields() if '__' in field]
            if related_fields:
                queryset = queryset.select_related(*set(related_fields))
            
            data = list(queryset)
            field_labels = self.get_field_labels()
            
            # Generate report based on format
            if format_type == 'csv':
                content = self.generate_csv(data, field_labels)
                response = HttpResponse(content, content_type='text/csv')
                response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
                
            elif format_type == 'excel':
                content = self.generate_excel(data, field_labels)
                response = HttpResponse(
                    content, 
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
                
            elif format_type == 'json':
                content = self.generate_json(data, field_labels)
                response = HttpResponse(content, content_type='application/json')
                response['Content-Disposition'] = f'attachment; filename="{filename}.json"'
                
            else:
                return Response(
                    {'error': 'Invalid format. Supported formats: csv, excel, json'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            return response
            
        except Exception as e:
            return Response(
                {'error': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

#     else:
#         return HttpResponse("Invalid format. Use ?format=pdf or ?format=excel", status=400)

# Example usage - create a specific report view
class OrderReportView(GenericReportDownloadView):
    """Example: User report download view."""
    permission_classes = [IsAdminUser]
    from django.contrib.auth.models import User  # Replace with your User model
    
    model = User
    fields = [
        'id',
        'username', 
        'email',
        'first_name',
        'last_name',
        'is_active',
        'date_joined',
        'last_login'
    ]
    field_labels = {
        'id': 'ID',
        'username': 'Username',
        'email': 'Email Address',
        'first_name': 'First Name',
        'last_name': 'Last Name',
        'is_active': 'Active Status',
        'date_joined': 'Registration Date',
        'last_login': 'Last Login Date'
    }
    
    def get_queryset(self):
        # Add any custom filtering logic here
        return super().get_queryset().filter(is_staff=False)


# Example with related fields
class UserReportView(GenericReportDownloadView):
    """Example: Order report with related fields."""
    permission_classes = [IsAdminUser]
    # Assuming you have an Order model with foreign keys
    model = User
    fields = [
        "first_name", "last_name", "email", "mobile", 'adhar_no', 'dob',
            'pan_no', 'zipcode', 'country', 'district__name', 'taluka__name', 'Village__name',
            'address', 'house_or_building', 'road_or_area', 'landmark',
            "message","state", "city","role"
    ]
    field_labels = {
    "first_name":       "First Name",
    "last_name":        "Last Name",
    "email":            "Email",
    "mobile":           "Mobile",
    "adhar_no":         "Aadhar Number",
    "dob":              "Date of Birth",
    "pan_no":           "PAN Number",
    "zipcode":          "Zip Code",
    "country":          "Country",
    "state":            "State",
    "district__name":   "District",
    "taluka__name":     "Taluka",
    "city":             "City",
    "Village__name":    "Village",
    "address":          "Address",
    "house_or_building":"House/Building",
    "road_or_area":     "Road/Area",
    "landmark":         "Landmark",
    "message":          "Message",
    "role" : "Role"
    }

class DistrictReportView(GenericReportDownloadView):
    """Example: Order report with related fields."""
    permission_classes = [IsAdminUser]

    # Assuming you have an Order model with foreign keys
    model = DistrictModel
    fields = [
        "id", "name",
    ]
    field_labels = {
    "id":     "ID",
    "name":       "District  Name"
    
    
    }


class TalukaReportView(GenericReportDownloadView):
    """Example: Order report with related fields."""
    permission_classes = [IsAdminUser]
    # Assuming you have an Order model with foreign keys
    model = TalukaModel
    fields = [
        "id", "name", "district__name"
    ]
    field_labels = {
    "id":     "ID",
    "name":       "taluka Name",
    "district__name" :" District Name"

    }

class VillageReportView(GenericReportDownloadView):
    """Example: Order report with related fields."""
    permission_classes = [IsAdminUser]
    # Assuming you have an Order model with foreign keys
    model = VillageModel
    fields = [
        "id", "name", "taluka__name","district__name"
    ]
    field_labels = {
    "id":     "ID",
    "name":       "Village",
    "taluka__name":  "Taluka",
    "district__name" :" District"
    
    }


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from users.models import User
from rest_framework import generics
from admin_panel.permissions import IsAdminUser
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from django.db.models import Count, Q
from django.contrib.auth import get_user_model
from vendor.models import *



class UserStatisticsListView(generics.ListAPIView):
    """
    Alternative approach using ListAPIView
    """

    permission_classes = [IsAdminUser]

    def list(self, request, *args, **kwargs):
        try:
            roles_stats = []

            # Get unique roles from database
            roles = User.objects.values_list("role", flat=True).distinct()

            for role in roles:
                count = User.objects.filter(role=role).count()
                roles_stats.append({"role": role, "count": count})

            total_users = User.objects.count()

            return Response(
                {
                    "success": True,
                    "message": "User statistics retrieved successfully.",
                    "data": {"total_users": total_users, "role_breakdown": roles_stats},
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": "An error occurred while retrieving user statistics.",
                    "error": str(e),
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )





class VehicleStatisticsView(APIView):
    """
    Generic API view to get vehicle statistics:
    - Vehicles per user/member
    - Total vehicles registered by type
    """

    permission_classes = [IsAuthenticated]  # Remove if public access needed

    def get(self, request, *args, **kwargs):
        try:
            # Get vehicles per user (only users who have vehicles)
            vehicles_per_user = (
                Vehicle.objects.values(
                    "user__id", "user__first_name", "user__last_name"
                )
                .annotate(vehicle_count=Count("id"))
                .order_by("-vehicle_count")
            )

            # Get total vehicles by type
            vehicles_by_type = (
                Vehicle.objects.values(
                    "vehicle_type__id",
                    "vehicle_type__name",  # Assuming VehicleType has a 'name' field
                )
                .annotate(total_vehicles=Count("id"))
                .order_by("-total_vehicles")
            )

            # Get overall statistics
            total_vehicles = Vehicle.objects.count()
            total_users_with_vehicles = (
                Vehicle.objects.values("user").distinct().count()
            )
            total_vehicle_types = (
                Vehicle.objects.values("vehicle_type").distinct().count()
            )

            # Get rental availability stats
            # rental_available_count = Vehicle.objects.filter(
            #     rental_available=True
            # ).count()
            # rental_not_available_count = Vehicle.objects.filter(
            #     rental_available=False
            # ).count()

            return Response(
                {
                    "success": True,
                    "message": "Vehicle statistics retrieved successfully.",
                    "data": {
                        "overview": {
                            "total_vehicles": total_vehicles,
                            "total_users_with_vehicles": total_users_with_vehicles,
                            "total_vehicle_types": total_vehicle_types,
                            # "rental_available": rental_available_count,
                            # "rental_not_available": rental_not_available_count,
                        },
                        "vehicles_per_user": [
                            {
                                "user_id": item["user__id"],
                                "full_name": f"{item['user__first_name'] or ''} {item['user__last_name'] or ''}".strip(),
                                "vehicle_count": item["vehicle_count"],
                            }
                            for item in vehicles_per_user
                        ],
                        "vehicles_by_type": [
                            {
                                "vehicle_type_id": item["vehicle_type__id"],
                                "vehicle_type_name": item["vehicle_type__name"],
                                "total_vehicles": item["total_vehicles"],
                            }
                            for item in vehicles_by_type
                        ],
                    },
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": "An error occurred while retrieving vehicle statistics.",
                    "error": str(e),
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class VehiclesByUserView(APIView):
    """
    API view to get vehicles count for a specific user
    """

    permission_classes = [IsAuthenticated]

    def get(self, request, user_id=None, *args, **kwargs):
        try:
            # If user_id is provided, get stats for that user, otherwise current user
            target_user_id = user_id or request.user.id

            # Check if user exists
            try:
                user = User.objects.get(id=target_user_id)
            except User.DoesNotExist:
                return Response(
                    {"success": False, "message": "User not found."},
                    status=status.HTTP_404_NOT_FOUND,
                )

            # Get user's vehicles by type
            user_vehicles_by_type = (
                Vehicle.objects.filter(user_id=target_user_id)
                .values("vehicle_type__id", "vehicle_type__name")
                .annotate(vehicle_count=Count("id"))
                .order_by("-vehicle_count")
            )

            # Get total count for this user
            total_user_vehicles = Vehicle.objects.filter(user_id=target_user_id).count()

            # Get rental stats for this user
            # rental_available = Vehicle.objects.filter(
            #     user_id=target_user_id  #rental_available=True
            # ).count()

            return Response(
                {
                    "success": True,
                    "message": f"Vehicle statistics for user retrieved successfully.",
                    "data": {
                        "user_info": {
                            "user_id": user.id,
                            "username": user.username,
                            "full_name": user.get_full_name(),
                        },
                        "vehicle_summary": {
                            "total_vehicles": total_user_vehicles,
                            # "rental_available": rental_available,
                            # "rental_not_available": total_user_vehicles
                            # - rental_available,
                        },
                        "vehicles_by_type": [
                            {
                                "vehicle_type_id": item["vehicle_type__id"],
                                "vehicle_type_name": item["vehicle_type__name"],
                                "count": item["vehicle_count"],
                            }
                            for item in user_vehicles_by_type
                        ],
                    },
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": "An error occurred while retrieving user vehicle statistics.",
                    "error": str(e),
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class VehicleTypeStatsView(APIView):
    """
    API view to get detailed statistics for each vehicle type
    """

    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        try:
            # Get detailed stats by vehicle type
            vehicle_type_stats = (
                VehicleType.objects.annotate(
                    total_vehicles=Count("vehicle"),
                    # rental_available_count=Count(
                    #     "vehicle", filter=Q(vehicle__rental_available=True)
                    # ),
                    unique_owners=Count("vehicle__user", distinct=True),
                )
                .values(
                    "id",
                    "name",  # Assuming VehicleType has 'name' field
                    "total_vehicles",
                    # "rental_available_count",
                    "unique_owners",
                )
                .order_by("-total_vehicles")
            )

            return Response(
                {
                    "success": True,
                    "message": "Vehicle type statistics retrieved successfully.",
                    "data": {
                        "vehicle_types": [
                            {
                                "vehicle_type_id": item["id"],
                                "vehicle_type_name": item["name"],
                                "total_vehicles": item["total_vehicles"],
                                # "rental_available": item["rental_available_count"],
                                # "rental_not_available": item["total_vehicles"]
                                # - item["rental_available_count"],
                                "unique_owners": item["unique_owners"],
                            }
                            for item in vehicle_type_stats
                        ]
                    },
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": "An error occurred while retrieving vehicle type statistics.",
                    "error": str(e),
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )